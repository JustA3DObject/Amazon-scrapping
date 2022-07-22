import csv
import requests
import json
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

asin_list = []
country_list = []
url_list = []
content = []
wb = load_workbook("Amazon Scraping.xlsx")
ws = wb.active
col_asin = ws["C"]
col_country = ws["D"]


def make_asin_list():
    for index, item in enumerate(col_asin):
        asin_list.insert(index, item.value)

    for index, item in enumerate(asin_list):
        if type(item) == float:
            asin_list[index] = int(item)

    asin_list.pop(0)


def make_country_list():
    for index, item in enumerate(col_country):
        country_list.insert(index, item.value)

    country_list.pop(0)


make_asin_list()
make_country_list()


def make_url_list():
    i = 0
    for asin, country in (zip(asin_list, country_list)):
        url_list.insert(i, f"https://www.amazon.{country}/dp/{asin}")
        i = i+1


make_url_list()
# print(url_list)

header = ["Product", "Price", "Details", "Image URL"]
with open("AmazonScraperData.csv", 'w', newline='', encoding='UTF8') as f:
    writer = csv.writer(f)
    writer.writerow(header)

j = 0
while(j < 100):

    URL = url_list[j]
    j = j+1
    # headers from https://httpbin.org/get

    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.114 Safari/537.36",
               "X-Amzn-Trace-Id": "Root=1-62d8036d-2b173c1f2e4e7a416cc9e554", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
               "Accept-Encoding": "gzip, deflate, br", "Accept-Language": "en-GB", }

    page = requests.get(URL, headers=headers)

    if page.status_code == 200:
        soup1 = BeautifulSoup(page.content, "html.parser")
        soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

        try:
            title = soup2.find(id="productTitle").get_text()
            title = title.strip()
        except AttributeError:
            print("Attribute Error: Title not found.")
            title = "Title not found."
        try:
            price = soup2.find(class_="a-offscreen").get_text()
            price = price.strip()
        except AttributeError:
            print("Attribute Error: Price not found.")
            price = "Price not found"
        try:
            about = soup2.find(
                class_="a-unordered-list a-vertical a-spacing-mini").get_text()
            about = about.strip()
        except AttributeError:
            print("Attribute Error: Description not found.")
            about = "Description not found"
        try:
            img_url = soup2.find(id="landingImage").get('src')
        except AttributeError:
            print("Attribute Error: Image not found.")
            img_url = "Image not found"

        else:
            json_list = {
                'title': title,
                'price': price,
                'about': about,
                'img_url': img_url
            }
            content.append(json_list)

            data = [title, price, about, img_url]
            with open("AmazonScraperData.csv", 'a+', newline='', encoding='UTF8') as f:
                writer = csv.writer(f)
                writer.writerow(data)

    else:
        print("404 ERROR!")
        title = "404 ERROR!"

        json_list = {
            'title': title,
            'price': '',
            'about': '',
            'img_url': ''
        }
        content.append(json_list)

        data = [title]
        with open("AmazonScraperData.csv", 'a+', newline='', encoding='UTF8') as f:
            writer = csv.writer(f)
            writer.writerow(data)

with open('AmazonScraperData.json', 'w', encoding='utf-8') as f:
    json.dump(content, f, ensure_ascii=False, indent=4)
